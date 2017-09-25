#!/usr/bin/env python
# -*- coding: utf-8 -*-

' openpyxl_test.py '

__author__ = 'Hyman Lee'


############## main code ###############

from openpyxl import Workbook
from openpyxl import load_workbook
import os

wb = Workbook()

def getwb(file_path):
    if not os.path.exists(file_path):
        return Workbook()
    return load_workbook(file_path)

def test(path, file):
    try:
        os.makedirs(path)
    except OSError:
        if not os.path.isdir(path):
            raise IOError, 'path (%s) is not a dir' % path
    xls_file = os.path.join(path, file)
    print xls_file
    if os.path.exists(xls_file) and not os.path.isfile(xls_file):
        raise Exception, '(%s) is not valid file path' % xls_file
    wb = getwb(xls_file)
    ws = wb.active
    headers = ['Name', 'zh-CHS', 'EN']
    if ws['A1'].value == None:
        # ws['A1'] = 'name'
        # ws['B1'] = 'zh-CHS'
        # ws['C1'] = 'EN'
         for col in range(len(headers)):
            ws.cell(row = 1, column = col + 1).value = headers[col]
    ws.append(['A', 'B', 'C'])
    wb.save(xls_file)




if __name__ == '__main__':
    test('./', 'test.xlsx')
