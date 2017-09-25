#!/usr/bin/env python
# -*- coding: utf-8 -*-

' excel_helper.py '

__author__ = 'Hyman Lee'


############## main code ###############

from openpyxl import Workbook
from openpyxl import load_workbook
import os
import const
import datetime

class ExcelHelper(object):
    """Excel 文件操作类"""
    def __init__(self, path, file):
        super(ExcelHelper, self).__init__()
        self.__getxlspath(path, file)

    def __getxlspath(self, path, file):
        if file.index('.')  <= 0:
            raise Exception, 'file (%s) is not invalid' % file
        file_strs = file.split('.')
        if len(file_strs) < 2 or file_strs[1] != 'xlsx':
            raise Exception, 'file (%s) is not invalid is not (.xlsx) type file' % file
        try:
            os.makedirs(path)
        except OSError:
            if not os.path.isdir(path):
                raise IOError, 'path (%s) is not a dir' % path
        xls_file = os.path.join(path, file)
        if os.path.exists(xls_file) and not os.path.isfile(xls_file):
            raise Exception, '(%s) is not valid file name' % file
        self.__xls_file = xls_file

    def __getwb(self):
        if not os.path.exists(self.__xls_file):
            return Workbook()
        return load_workbook(self.__xls_file)

    def record(self, translated_datas, ):
        wb = self.__getwb()
        ws = wb.active
        headers = ['Name', const.SOURCE, const.TARGET, 'Editor', 'Date']
        if ws['A1'].value == None:
            for col in range(len(headers)):
                ws.cell(row = 1, column = col + 1).value = headers[col]
        edit_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for name in translated_datas:
            ws.append([name, translated_datas[name][const.SOURCE], translated_datas[name][const.TARGET], const.EDITOR, edit_time])
        wb.save(self.__xls_file)


if __name__ == '__main__':
    eh = ExcelHelper('./', 'test.xlsx')

